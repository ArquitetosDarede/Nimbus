"""
XLSX Analysis Agent - Specialized agent for XLSX file analysis
"""

import logging
from typing import Dict, Any
from pathlib import Path

logger = logging.getLogger(__name__)

# Optional imports
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class XlsxAnalysisAgent:
    """Specialized agent for XLSX file analysis."""

    def analyze(self, file_path: Path) -> Dict[str, Any]:
        """Extract data from XLSX files."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl not available for XLSX extraction")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheets_data = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []

            # Get all rows
            for row in sheet.iter_rows(values_only=True):
                # Convert None to empty string and filter out completely empty rows
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                if any(cleaned_row):  # Only add non-empty rows
                    sheet_data.append(cleaned_row)

            if sheet_data:
                # First row as headers if it exists
                headers = sheet_data[0] if sheet_data else []
                sheets_data.append({
                    "sheet_name": sheet_name,
                    "data": sheet_data,
                    "headers": headers,
                    "row_count": len(sheet_data)
                })

        # Convert to text format for analysis
        text_content = []
        for sheet_info in sheets_data:
            text_content.append(f"Sheet: {sheet_info['sheet_name']}")
            if sheet_info['headers']:
                text_content.append("Headers: " + ", ".join(sheet_info['headers']))
            text_content.append("Data:")
            for row in sheet_info['data'][:10]:  # Limit to first 10 rows for text
                text_content.append(", ".join(row))
            if len(sheet_info['data']) > 10:
                text_content.append(f"... and {len(sheet_info['data']) - 10} more rows")

        return {
            "extracted_text": "\n\n".join(text_content),
            "extracted_data": {
                "sheets": sheets_data,
                "metadata": {
                    "total_sheets": len(sheets_data),
                    "total_rows": sum(sheet['row_count'] for sheet in sheets_data)
                }
            }
        }